"""
svg_layout_parser.py - Parse Visio-exported SVG layout drawings into
structured JSON for SeaTec3D 3D scene generation.

Handles:
  - rect          -> box (skid, package, exchanger)
  - ellipse/circle -> vertical_vessel
  - path with arcs -> horizontal_vessel (stadium/capsule plan view)
  - Dimension annotations (arrows + text) for scale calibration
  - Grid line labels (A, B, 1, 2 etc.)
  - Equipment tags from <desc>, <text>, or Visio custom properties
  - Coordinate origin at bottom-left of module boundary, Y-up

Usage:
  parser = SvgLayoutParser("Example01.svg")
  result = parser.parse()
  parser.save_json("output.json")
"""

import xml.etree.ElementTree as ET
import json
import re
import math
import sys
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Tuple


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
NS = {
    "svg": "http://www.w3.org/2000/svg",
    "v": "http://schemas.microsoft.com/visio/2003/SVGExtensions/",
    "xlink": "http://www.w3.org/1999/xlink",
}

DEFAULT_HEIGHT_MM = 2000
DEFAULT_WEIGHT_KG = 1000
DEFAULT_ELEVATION_MM = 0


# Excel header aliases (all matched case-insensitively after .strip().lower())
_EXCEL_HEADER_ALIASES = {
    "tag": [
        "tag", "equipment tag", "equipment_tag", "item",
        "tag number", "tag_number",
    ],
    "height_mm": [
        "height_mm", "height", "height (mm)", "h_mm", "tan-tan",
    ],
    "weight_kg": [
        "weight_kg", "weight", "weight (kg)",
        "dry weight", "operating weight",
        "dry weight (kg)", "operating weight (kg)",
        "weight (t)", "weight (mt)",
        "dry weight (t)", "operating weight (t)",
        "dry weight (mt)", "operating weight (mt)",
    ],
    "diameter_mm": [
        "diameter_mm", "diameter", "dia", "od", "dia (mm)", "id",
    ],
    "length_mm": [
        "length_mm", "length", "len", "length (mm)", "t/t", "tan to tan",
    ],
    "elevation_mm": [
        "elevation_mm", "elevation", "elev", "el", "elevation (mm)",
    ],
}

# Marker arrow setback (pts) - standard Visio dimension arrow markers
MARKER_SETBACK_START = 9.33   # refX * strokeWidth for start marker
MARKER_SETBACK_END = 9.42     # refX * strokeWidth for end marker
MARKER_TOTAL = MARKER_SETBACK_START + MARKER_SETBACK_END  # 18.75 pts


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
def looks_like_equipment_tag(text: str) -> bool:
    """Return True if `text` looks like an equipment / feature tag.

    Broader than the original `[A-Za-z]{1,4}-\\d` rule so tags such as
    `WH-A-1201`, `GAS-LIFT`, `FIRE-WALL`, `HATCH`, `WELLBAY-AREA` are
    recognised alongside `K-3170-A` style names. Pure numeric strings
    (dimension values like `24150`) are rejected so they remain
    classified as dim values.
    """
    if not text:
        return False
    s = text.strip()
    if len(s) < 3:
        return False
    if s.isdigit():
        return False
    # Must contain at least one letter; reject things like "-123".
    for ch in s:
        if ch.isalpha():
            return True
    return False


@dataclass
class SvgShape:
    """Raw shape extracted from SVG before calibration."""
    svg_id: str
    shape_type: str           # rect, ellipse, stadium, path_other
    title: str = ""           # Visio shape title
    visio_class: str = ""     # ShapeClass from v:ud
    visio_subtype: str = ""   # SubType value
    tag: str = ""             # equipment tag
    cx_pts: float = 0.0       # center x in SVG pts
    cy_pts: float = 0.0       # center y in SVG pts
    width_pts: float = 0.0    # bounding width
    height_pts: float = 0.0   # bounding height
    rx_pts: float = 0.0       # radius x (ellipse)
    ry_pts: float = 0.0       # radius y (ellipse)
    rotation_deg: float = 0.0
    css_class: str = ""


@dataclass
class DimAnnotation:
    """Dimension annotation: arrow span + value label."""
    svg_id: str
    value_mm: int
    direction: str = ""       # horizontal or vertical
    span_pts: float = 0.0     # corrected span including marker setback
    start_abs: float = 0.0    # absolute start position (x or y)
    end_abs: float = 0.0      # absolute end position
    perp_abs: float = 0.0     # perpendicular SVG coord (svg_y for horiz,
                              # svg_x for vertical) used to preserve the
                              # original stack position of the annotation
    label_x_pts: float = 0.0  # label group translate x (for tiebreaker)
    label_y_pts: float = 0.0  # label group translate y (for tiebreaker)
    label_rot: float = 0.0    # label rotation (90/-90 => vertical dim)


@dataclass
class GridLabel:
    """Grid/column line label (A, B, 1, 2, etc.)."""
    label: str
    x_pts: float
    y_pts: float


# ---------------------------------------------------------------------------
# Transform parser
# ---------------------------------------------------------------------------
def parse_transform(t: str) -> Tuple[float, float, float]:
    """Extract translate(tx,ty) and rotate(deg) from SVG transform string."""
    tx, ty, rot = 0.0, 0.0, 0.0
    m = re.search(r"translate\(([-\d.]+)\s*,\s*([-\d.]+)\)", t or "")
    if m:
        tx, ty = float(m.group(1)), float(m.group(2))
    m = re.search(r"rotate\(([-\d.]+)\)", t or "")
    if m:
        rot = float(m.group(1))
    return tx, ty, rot


def apply_rotation(px, py, rot_deg, origin_x=0, origin_y=0):
    """Rotate point (px,py) by rot_deg around origin."""
    if rot_deg == 0:
        return px, py
    rad = math.radians(rot_deg)
    dx, dy = px - origin_x, py - origin_y
    rx = dx * math.cos(rad) - dy * math.sin(rad) + origin_x
    ry = dx * math.sin(rad) + dy * math.cos(rad) + origin_y
    return rx, ry


# ---------------------------------------------------------------------------
# Path analysis for stadium/capsule detection
# ---------------------------------------------------------------------------
def parse_path_d(d: str) -> list:
    """Tokenize an SVG path d attribute into command segments."""
    # Split on command letters, keeping the letter
    tokens = re.findall(r"[MmLlHhVvCcSsQqTtAaZz][^MmLlHhVvCcSsQqTtAaZz]*", d)
    return tokens


def is_stadium_path(d: str) -> Optional[dict]:
    """
    Detect stadium/capsule shape: 2 arcs + 2 lines (or close path).
    Returns dict with local bounding box if stadium, else None.

    Typical Visio vessel path:
      M x0 y0
      A rx ry rot large-arc sweep x1 y1   (end cap)
      L x2 y2                              (side)
      A rx ry rot large-arc sweep x3 y3   (other end cap)
      L x4 y4  (or Z)                     (close)
    """
    tokens = parse_path_d(d.strip())
    arcs = []
    lines = []
    moves = []
    coords = []

    for tok in tokens:
        cmd = tok[0]
        nums = [float(x) for x in re.findall(r"[-+]?\d*\.?\d+", tok[1:])]
        if cmd in ("M", "m"):
            if len(nums) >= 2:
                moves.append((nums[0], nums[1]))
                coords.append((nums[0], nums[1]))
        elif cmd in ("L", "l"):
            if len(nums) >= 2:
                lines.append((nums[0], nums[1]))
                coords.append((nums[0], nums[1]))
        elif cmd in ("A", "a"):
            # A rx ry x-rotation large-arc-flag sweep-flag x y
            if len(nums) >= 7:
                arcs.append({
                    "rx": nums[0], "ry": nums[1],
                    "rotation": nums[2],
                    "large_arc": int(nums[3]),
                    "sweep": int(nums[4]),
                    "x": nums[5], "y": nums[6]
                })
                coords.append((nums[5], nums[6]))
        elif cmd in ("Z", "z"):
            pass

    # Stadium: exactly 2 arcs + 1 or 2 line segments
    if len(arcs) == 2 and len(lines) >= 1:
        all_x = [c[0] for c in coords]
        all_y = [c[1] for c in coords]
        x_min, x_max = min(all_x), max(all_x)
        y_min, y_max = min(all_y), max(all_y)

        return {
            "x_min": x_min, "x_max": x_max,
            "y_min": y_min, "y_max": y_max,
            "width": x_max - x_min,
            "height": y_max - y_min,
            "arc_rx": arcs[0]["rx"],
            "arc_ry": arcs[0]["ry"],
        }

    return None


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------
class SvgLayoutParser:
    def __init__(self, svg_path: str = None, excel_path: str = None):
        self.svg_path = svg_path
        self.excel_path = excel_path
        self.warnings: List[str] = []

        # Standalone Excel-merge mode: no SVG to parse
        if svg_path is None:
            self.tree = None
            self.root = None
            self.shapes = []
            self.dim_arrows = {}
            self.dim_labels = []
            self.grid_labels = []
            self.label_texts = []
            self.scale = None
            self.origin_x = 0.0
            self.origin_y_bottom = 0.0
            self.module_width_mm = None
            self.module_length_mm = None
            self.page_w = 0.0
            self.page_h = 0.0
            self.css_fills = {}
            self.page_props = {}
            return

        self.tree = ET.parse(svg_path)
        self.root = self.tree.getroot()

        # Extracted raw data
        self.shapes: List[SvgShape] = []
        self.dim_arrows: dict = {}       # svg_id -> path span info
        self.dim_labels: List[DimAnnotation] = []
        self.grid_labels: List[GridLabel] = []
        self.label_texts: list = []       # (svg_id, text, tx, ty, local_cx, local_cy)

        # Calibration
        self.scale = None                 # mm per SVG pt
        self.origin_x = 0.0              # SVG x of left boundary
        self.origin_y_bottom = 0.0       # SVG y of bottom boundary
        self.module_width_mm = None
        self.module_length_mm = None

        # Page metrics
        vb = self.root.get("viewBox", "0 0 841.89 595.276")
        parts = vb.split()
        self.page_w = float(parts[2])
        self.page_h = float(parts[3])

        # CSS class analysis - find which classes are equipment vs annotation
        self._parse_css_classes()

        # Visio page properties (drawingScale, pageScale, drawingUnits)
        self._extract_page_properties()

    # -------------------------------------------------------------------
    # CSS class extraction
    # -------------------------------------------------------------------
    def _parse_css_classes(self):
        """Parse <style> to identify fill classes."""
        self.css_fills = {}
        style_el = self.root.find(".//svg:style", NS)
        if style_el is not None and style_el.text:
            for m in re.finditer(
                r"\.(st\d+)\s*\{([^}]+)\}", style_el.text
            ):
                cls_name = m.group(1)
                props = m.group(2)
                fill_m = re.search(r"fill:\s*(#[0-9a-fA-F]+|none)", props)
                stroke_m = re.search(r"stroke:\s*(#[0-9a-fA-F]+|none)", props)
                marker_m = re.search(r"marker-", props)
                dash_m = re.search(r"stroke-dasharray", props)
                self.css_fills[cls_name] = {
                    "fill": fill_m.group(1) if fill_m else None,
                    "stroke": stroke_m.group(1) if stroke_m else None,
                    "is_marker": marker_m is not None,
                    "is_dashed": dash_m is not None,
                }

    def _is_equipment_class(self, cls: str) -> bool:
        """Check if a CSS class represents equipment (filled, not annotation).

        Accepts any non-empty fill that isn't pure white or pure black.
        Callers also fall back to Visio metadata (ShapeClass=Equipment or
        recognized title) when this returns False -- some Visio templates
        export equipment with white fill (e.g. Example02 stadiums on st1).
        """
        info = self.css_fills.get(cls, {})
        fill = info.get("fill")
        if fill and fill != "none" and fill != "#ffffff" and fill != "#000000":
            return True
        return False

    def _shape_is_equipment(self, cls: str, visio_class: str,
                            title: str) -> bool:
        """Combined equipment test: CSS fill OR Visio metadata."""
        if cls and self._is_equipment_class(cls):
            return True
        if visio_class == "Equipment":
            return True
        if title and title.lower() in (
            "vessel", "drum", "separator", "tank", "column",
            "compressor", "pump", "exchanger",
        ):
            return True
        return False

    def _extract_page_properties(self):
        """Read Visio v:pageProperties for fallback scale calibration."""
        self.page_props = {}
        if self.root is None:
            return
        pp = self.root.find(".//v:pageProperties", NS)
        if pp is None:
            return
        for attr in ("drawingScale", "pageScale", "drawingUnits"):
            v = pp.get("{%s}%s" % (NS["v"], attr))
            if v is None:
                continue
            try:
                self.page_props[attr] = float(v)
            except (TypeError, ValueError):
                pass

    def _is_dimension_arrow_class(self, cls: str) -> bool:
        info = self.css_fills.get(cls, {})
        return info.get("is_marker", False)

    def _is_boundary_class(self, cls: str) -> bool:
        info = self.css_fills.get(cls, {})
        return info.get("is_dashed", False)

    # -------------------------------------------------------------------
    # Shape extraction
    # -------------------------------------------------------------------
    def _extract_shapes(self):
        """Walk all <g> elements and classify shapes."""
        for g in self.root.findall(".//svg:g[@id]", NS):
            svg_id = g.get("id", "")
            if not svg_id.startswith("shape"):
                continue

            transform = g.get("transform", "")
            tx, ty, rot = parse_transform(transform)

            title_el = g.find("svg:title", NS)
            title = title_el.text.strip() if title_el is not None and title_el.text else ""

            desc_el = g.find("svg:desc", NS)
            text_el = g.find(".//svg:text", NS)
            text_content = "".join(text_el.itertext()).strip() if text_el is not None else ""

            visio_class = ""
            visio_subtype = ""
            tag = ""

            ud_nodes = list(g.findall(".//v:ud", NS))
            for ud in ud_nodes:
                name = ud.get("{%s}nameU" % NS["v"], "")
                val = ud.get("{%s}val" % NS["v"], "")

                if name == "ShapeClass":
                    visio_class = val.replace("VT4(", "").rstrip(")")
                elif name == "SubType" and "VT4(" in (val or ""):
                    visio_subtype = val.replace("VT4(", "").rstrip(")")
                elif name == "PEComponentTag" and val:
                    clean = val.replace("VT4(", "").rstrip(")")
                    if clean and clean != "E-0":
                        tag = clean

            if not tag and desc_el is not None and desc_el.text:
                txt = desc_el.text.strip()
                if looks_like_equipment_tag(txt):
                    tag = txt

            group_added_shape = False

            # 1. Rectangle
            for rect in g.findall(".//svg:rect", NS):
                cls = rect.get("class", "")
                if self._shape_is_equipment(cls, visio_class, title):
                    rx = float(rect.get("x", 0))
                    ry_local = float(rect.get("y", 0))
                    rw = float(rect.get("width", 0))
                    rh = float(rect.get("height", 0))
                    cx_local = rx + rw / 2
                    cy_local = ry_local + rh / 2
                    cx_abs, cy_abs = cx_local + tx, cy_local + ty

                    if rot != 0:
                        cx_abs, cy_abs = apply_rotation(cx_local, cy_local, rot)
                        cx_abs += tx
                        cy_abs += ty

                    self.shapes.append(SvgShape(
                        svg_id=svg_id,
                        shape_type="rect",
                        title=title,
                        visio_class=visio_class,
                        visio_subtype=visio_subtype,
                        tag=tag,
                        cx_pts=cx_abs,
                        cy_pts=cy_abs,
                        width_pts=rw,
                        height_pts=rh,
                        rotation_deg=rot,
                        css_class=cls,
                    ))
                    group_added_shape = True
                    break

            # 2. Ellipse / Circle
            if not group_added_shape:
                for ell in g.findall(".//svg:ellipse", NS):
                    cls = ell.get("class", "")
                    if self._shape_is_equipment(cls, visio_class, title):
                        ecx = float(ell.get("cx", 0))
                        ecy = float(ell.get("cy", 0))
                        erx = float(ell.get("rx", 0))
                        ery = float(ell.get("ry", 0))
                        cx_abs = tx + ecx
                        cy_abs = ty + ecy

                        self.shapes.append(SvgShape(
                            svg_id=svg_id,
                            shape_type="ellipse",
                            title=title,
                            visio_class=visio_class,
                            visio_subtype=visio_subtype,
                            tag=tag,
                            cx_pts=cx_abs,
                            cy_pts=cy_abs,
                            rx_pts=erx,
                            ry_pts=ery,
                            rotation_deg=rot,
                            css_class=cls,
                        ))
                        group_added_shape = True
                        break

            if not group_added_shape:
                for circ in g.findall(".//svg:circle", NS):
                    cls = circ.get("class", "")
                    if self._shape_is_equipment(cls, visio_class, title):
                        ccx = float(circ.get("cx", 0))
                        ccy = float(circ.get("cy", 0))
                        cr = float(circ.get("r", 0))
                        cx_abs = tx + ccx
                        cy_abs = ty + ccy

                        self.shapes.append(SvgShape(
                            svg_id=svg_id,
                            shape_type="ellipse",
                            title=title,
                            visio_class=visio_class,
                            visio_subtype=visio_subtype,
                            tag=tag,
                            cx_pts=cx_abs,
                            cy_pts=cy_abs,
                            rx_pts=cr,
                            ry_pts=cr,
                            rotation_deg=rot,
                            css_class=cls,
                        ))
                        group_added_shape = True
                        break

            # 3. Paths
            for path in g.findall(".//svg:path", NS):
                cls = path.get("class", "")
                d = path.get("d", "")
                if not d:
                    continue

                stadium = is_stadium_path(d)
                if stadium and (visio_class == "Equipment" or
                                title in ("Vessel", "Column", "Tank",
                                        "Drum", "Separator")):
                    local_cx = (stadium["x_min"] + stadium["x_max"]) / 2
                    local_cy = (stadium["y_min"] + stadium["y_max"]) / 2
                    w = stadium["width"]
                    h = stadium["height"]

                    if rot != 0:
                        cx_abs, cy_abs = apply_rotation(local_cx, local_cy, rot)
                        cx_abs += tx
                        cy_abs += ty
                        if abs(rot) in (90, -90, 270):
                            w, h = h, w
                    else:
                        cx_abs = tx + local_cx
                        cy_abs = ty + local_cy

                    self.shapes.append(SvgShape(
                        svg_id=svg_id,
                        shape_type="stadium",
                        title=title,
                        visio_class=visio_class,
                        visio_subtype=visio_subtype,
                        tag=tag,
                        cx_pts=cx_abs,
                        cy_pts=cy_abs,
                        width_pts=w,
                        height_pts=h,
                        rotation_deg=rot,
                        css_class=cls,
                    ))
                    group_added_shape = True
                    continue

                if self._is_dimension_arrow_class(cls):
                    nums = [float(x) for x in re.findall(r"[-+]?\d*\.?\d+", d)]
                    if len(nums) >= 4:
                        xs = nums[0::2]
                        ys = nums[1::2]
                        x_span = max(xs) - min(xs)
                        y_span = max(ys) - min(ys)
                        span = max(x_span, y_span)
                        corrected_span = span + MARKER_TOTAL

                        if abs(rot) < 1:
                            start = tx + min(xs) - MARKER_SETBACK_START
                            end = tx + max(xs) + MARKER_SETBACK_END
                            perp = ty + ys[0]
                            self.dim_arrows[svg_id] = {
                                "direction": "horizontal",
                                "span_pts": corrected_span,
                                "start_abs": start,
                                "end_abs": end,
                                "perp_abs": perp,
                            }
                        else:
                            if rot > 0:
                                y_min = ty + min(xs) - MARKER_SETBACK_START
                                y_max = ty + max(xs) + MARKER_SETBACK_END
                                perp = tx - ys[0]
                            else:
                                y_min = ty - max(xs) - MARKER_SETBACK_END
                                y_max = ty - min(xs) + MARKER_SETBACK_START
                                perp = tx + ys[0]

                            self.dim_arrows[svg_id] = {
                                "direction": "vertical",
                                "span_pts": corrected_span,
                                "start_abs": y_min,
                                "end_abs": y_max,
                                "perp_abs": perp,
                            }
                    continue

                if self._is_boundary_class(cls):
                    continue

            best_text = tag or text_content
            dim_value = None

            if best_text and re.match(r"^\d+$", best_text):
                dim_value = int(best_text)
                best_text = None

            if best_text and looks_like_equipment_tag(best_text) and not group_added_shape:
                text_rect = g.find(".//v:textRect", NS)
                lcx, lcy = 0, 0
                if text_rect is not None:
                    lcx = float(text_rect.get("cx", 0))
                    lcy = float(text_rect.get("cy", 0))
                if rot != 0:
                    lcx, lcy = apply_rotation(lcx, lcy, rot)

                self.label_texts.append((svg_id, best_text, tx, ty, lcx, lcy))

            if dim_value is not None:
                self.dim_labels.append(DimAnnotation(
                    svg_id=svg_id,
                    value_mm=dim_value,
                    label_x_pts=tx,
                    label_y_pts=ty,
                    label_rot=rot,
                ))

            if (text_content and
                    re.match(r"^[A-Z0-9]$", text_content) and
                    not tag):
                text_rect = g.find(".//v:textRect", NS)
                lcx, lcy = 0, 0
                if text_rect is not None:
                    lcx = float(text_rect.get("cx", 0))
                    lcy = float(text_rect.get("cy", 0))
                abs_x = tx + lcx
                abs_y = ty + lcy
                self.grid_labels.append(GridLabel(
                    label=text_content,
                    x_pts=abs_x,
                    y_pts=abs_y,
                ))

    # -------------------------------------------------------------------
    # Label matching
    # -------------------------------------------------------------------
    def _match_labels_to_shapes(self):
        """Associate text labels with nearest equipment shape."""
        unlabeled = [s for s in self.shapes if not s.tag]

        for _sid, text, tx, ty, lcx, lcy in self.label_texts:
            abs_x = tx + lcx
            abs_y = ty + lcy

            best = None
            best_d = 1e9

            for s in unlabeled:
                d = math.hypot(s.cx_pts - abs_x, s.cy_pts - abs_y)
                if d < best_d:
                    best_d = d
                    best = s

            if best and best_d < 500:
                best.tag = text
                try:
                    unlabeled.remove(best)
                except ValueError:
                    pass

        idx = 1
        for s in self.shapes:
            if not s.tag:
                prefix = "EQ"
                if s.title:
                    prefix = s.title[:3].upper()
                s.tag = "%s-%03d" % (prefix, idx)
                self.warnings.append(
                    "No tag found for %s; assigned '%s'" % (s.svg_id, s.tag)
                )
                idx += 1

    # -------------------------------------------------------------------
    # Dimension pairing: match value labels to arrows
    # -------------------------------------------------------------------
    def _pair_dimensions(self):
        """Match dimension value labels to arrows by comparing arrow length
        (in mm) against the label's text value.

        Uses a bootstrap scale from Visio page properties to convert each
        arrow's span_pts into mm, then pairs each label to the arrow whose
        implied length is closest to the label's number (within 3%). Ties
        are broken by geometric proximity in SVG space. Falls back to the
        legacy SVG-ID-proximity heuristic when no bootstrap scale is
        available (Visio exports lacking v:pageProperties).
        """
        bootstrap = self._bootstrap_scale()
        if bootstrap is None:
            self._pair_dimensions_by_id()
            return

        TOL = 0.03  # 3% length tolerance
        paired = []
        used_arrows = set()

        # Process labels large-first so dominant dims claim their arrows
        # before any smaller ones that could land in the tolerance band.
        # Note: label rotation is NOT used to filter direction -- Visio
        # often draws labels horizontally even for vertical arrows, so
        # we match purely on measured length and disambiguate by position.
        for dl in sorted(self.dim_labels, key=lambda d: -d.value_mm):
            if dl.value_mm <= 0:
                continue
            candidates = []
            for aid, arrow in self.dim_arrows.items():
                if aid in used_arrows:
                    continue
                implied_mm = arrow["span_pts"] * bootstrap
                rel_err = abs(implied_mm - dl.value_mm) / dl.value_mm
                if rel_err < TOL:
                    candidates.append((aid, arrow, rel_err))

            if not candidates:
                self.warnings.append(
                    "Dimension label '%d' has no matching arrow "
                    "(bootstrap scale %.3f mm/pt); label skipped"
                    % (dl.value_mm, bootstrap)
                )
                continue

            if len(candidates) == 1:
                aid, arrow, _ = candidates[0]
            else:
                # Tiebreaker: nearest arrow midpoint to label position
                lx, ly = dl.label_x_pts, dl.label_y_pts

                def _dist(c):
                    arrow = c[1]
                    if arrow["direction"] == "horizontal":
                        mx = (arrow["start_abs"] + arrow["end_abs"]) / 2.0
                        my = arrow.get("perp_abs", 0.0)
                    else:
                        mx = arrow.get("perp_abs", 0.0)
                        my = (arrow["start_abs"] + arrow["end_abs"]) / 2.0
                    return math.hypot(mx - lx, my - ly)

                aid, arrow, _ = min(candidates, key=_dist)

            dl.direction = arrow["direction"]
            dl.span_pts = arrow["span_pts"]
            dl.start_abs = arrow["start_abs"]
            dl.end_abs = arrow["end_abs"]
            dl.perp_abs = arrow.get("perp_abs", 0.0)
            paired.append(dl)
            used_arrows.add(aid)

        self.dim_labels = paired

    def _bootstrap_scale(self):
        """Return a seed mm-per-pt scale for value-based arrow pairing.

        Uses Visio page properties when present; returns None otherwise
        so the caller can fall back to ID-proximity pairing.
        """
        ds = self.page_props.get("drawingScale")
        ps = self.page_props.get("pageScale")
        if ds and ps and ps != 0:
            return (ds / ps) * (25.4 / 72.0)
        return None

    def _pair_dimensions_by_id(self):
        """Legacy pairing: nearest SVG shape ID. Used only as a fallback
        when no bootstrap scale is available.
        """
        paired = []
        used_arrows = set()

        for dl in self.dim_labels:
            label_num = int(re.search(r"\d+", dl.svg_id).group())
            best_arrow_id = None
            best_dist = 999
            for aid in self.dim_arrows:
                if aid in used_arrows:
                    continue
                arrow_num = int(re.search(r"\d+", aid).group())
                dist = abs(arrow_num - label_num)
                if dist < best_dist:
                    best_dist = dist
                    best_arrow_id = aid

            if best_arrow_id and best_dist <= 3:
                arrow = self.dim_arrows[best_arrow_id]
                dl.direction = arrow["direction"]
                dl.span_pts = arrow["span_pts"]
                dl.start_abs = arrow["start_abs"]
                dl.end_abs = arrow["end_abs"]
                dl.perp_abs = arrow.get("perp_abs", 0.0)
                paired.append(dl)
                used_arrows.add(best_arrow_id)

        self.dim_labels = paired

    # -------------------------------------------------------------------
    # Scale calibration
    # -------------------------------------------------------------------
    def _calibrate_scale(self):
        """
        Determine mm-per-SVG-pt scale from dimension annotations,
        falling back to Visio page properties or scale=1.0.
        """
        ranked = sorted(self.dim_labels, key=lambda d: d.value_mm,
                        reverse=True) if self.dim_labels else []
        scales = [d.value_mm / d.span_pts for d in ranked if d.span_pts > 0]

        if not scales:
            # Fallback 1: Visio page properties
            ds = self.page_props.get("drawingScale")
            ps = self.page_props.get("pageScale")
            if ds and ps and ps != 0:
                # SVG: 72 pt/inch -> 25.4/72 mm per pt at 1:1.
                # Visio scale ratio (real per page) multiplies that.
                self.scale = (ds / ps) * (25.4 / 72.0)
                self.warnings.append(
                    "Scale calibrated from Visio page properties: "
                    "%.3f mm/pt (no dimension arrows present)" % self.scale
                )
                return
            # Fallback 2: identity
            self.warnings.append(
                "No scale calibration available; "
                "coordinates are in SVG points"
            )
            self.scale = 1.0
            return

        # Use the largest dimension's scale as primary
        self.scale = scales[0]

        # Check consistency across all dimensions
        if len(scales) > 1:
            max_dev = max(abs(s - self.scale) / self.scale for s in scales[1:])
            if max_dev > 0.05:
                self.warnings.append(
                    "Scale inconsistency detected (%.1f%% max deviation); "
                    "using largest dimension for calibration (%.4f mm/pt)"
                    % (max_dev * 100, self.scale)
                )

        # Identify module boundary dimensions
        for d in ranked:
            if d.direction == "horizontal" and self.module_width_mm is None:
                self.module_width_mm = d.value_mm
            elif d.direction == "vertical" and self.module_length_mm is None:
                self.module_length_mm = d.value_mm

    # -------------------------------------------------------------------
    # Origin detection
    # -------------------------------------------------------------------
    def _detect_origin(self):
        """
        Determine coordinate origin (bottom-left of module boundary).
        Uses the largest horizontal dimension arrow for X origin
        and largest vertical for Y origin.
        """
        # Find the module-width dimension (largest horizontal)
        h_dims = sorted(
            [d for d in self.dim_labels if d.direction == "horizontal"],
            key=lambda d: d.value_mm, reverse=True
        )
        v_dims = sorted(
            [d for d in self.dim_labels if d.direction == "vertical"],
            key=lambda d: d.value_mm, reverse=True
        )

        if h_dims:
            self.origin_x = h_dims[0].start_abs
        else:
            # Fallback: use leftmost equipment shape
            if self.shapes:
                self.origin_x = min(
                    s.cx_pts - s.width_pts / 2 for s in self.shapes
                ) - 100 * (1.0 / (self.scale or 1))

        if v_dims:
            # Bottom boundary = end of vertical span (highest SVG y = bottom)
            self.origin_y_bottom = v_dims[0].end_abs
        else:
            if self.shapes:
                max_y_vals = []
                for s in self.shapes:
                    half_h = s.height_pts / 2 if s.height_pts else s.ry_pts
                    max_y_vals.append(s.cy_pts + half_h)
                self.origin_y_bottom = max(max_y_vals) + 50
            else:
                self.origin_y_bottom = self.page_h

    # -------------------------------------------------------------------
    # Equipment type classification
    # -------------------------------------------------------------------
    def _classify_equipment(self, shape: SvgShape) -> str:
        """Determine 3D equipment type from SVG shape + metadata."""
        # Visio metadata takes priority
        title_lower = shape.title.lower()
        if title_lower in ("vessel", "drum", "separator", "tank"):
            if shape.shape_type == "stadium":
                return "horizontal_vessel"
            elif shape.shape_type == "ellipse":
                return "vertical_vessel"
        if title_lower == "column":
            if shape.shape_type == "stadium":
                return "horizontal_vessel"
            return "vertical_vessel"

        # Fall back to shape geometry
        if shape.shape_type == "rect":
            return "box"
        elif shape.shape_type == "ellipse":
            # Circle in plan = vertical vessel
            return "vertical_vessel"
        elif shape.shape_type == "stadium":
            return "horizontal_vessel"

        return "box"

    # -------------------------------------------------------------------
    # SVG -> real-world coordinate conversion
    # -------------------------------------------------------------------
    def _to_real(self, svg_x, svg_y):
        """Convert SVG pts to real-world mm (origin bottom-left, Y-up)."""
        real_x = (svg_x - self.origin_x) * self.scale
        real_y = (self.origin_y_bottom - svg_y) * self.scale
        return round(real_x), round(real_y)

    def _to_real_dim(self, pts):
        """Convert SVG dimension to mm."""
        return round(pts * self.scale)

    # -------------------------------------------------------------------
    # Excel equipment list merge
    # -------------------------------------------------------------------
    def _merge_excel(self, equipment_list: list):
        """
        Merge dimensional/weight data from Excel equipment list.
        Matches on equipment tag using flexible header aliases.
        Auto-converts tonnes->kg and metres->mm where unit hints are present.
        """
        if not self.excel_path:
            return

        try:
            import openpyxl

            wb = openpyxl.load_workbook(
                self.excel_path,
                data_only=True,
                read_only=True,
            )
            ws = wb.active

            col_idx_by_key = {}
            weight_header = ""

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), [])
            for col_idx, cell in enumerate(header_row, 1):
                if cell.value is None:
                    continue

                norm = str(cell.value).strip().lower()
                for canonical, aliases in _EXCEL_HEADER_ALIASES.items():
                    if canonical in col_idx_by_key:
                        continue
                    if norm in aliases:
                        col_idx_by_key[canonical] = col_idx
                        if canonical == "weight_kg":
                            weight_header = str(cell.value).strip()
                        break

            tag_col = col_idx_by_key.get("tag")
            if not tag_col:
                self.warnings.append(
                    "Excel file has no recognizable 'tag' column; skipping merge"
                )
                return

            wh_lower = weight_header.lower()
            weight_is_tonnes = ("(t)" in wh_lower) or ("(mt)" in wh_lower)

            numeric_keys = (
                "height_mm", "weight_kg", "diameter_mm",
                "length_mm", "elevation_mm",
            )

            excel_data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                tag_val = row[tag_col - 1] if tag_col - 1 < len(row) else None
                if not tag_val:
                    continue

                tag_val = str(tag_val).strip()
                entry = {}

                for key in numeric_keys:
                    col = col_idx_by_key.get(key)
                    if not col or col - 1 >= len(row):
                        continue

                    v = row[col - 1]
                    if v is None or v == "":
                        continue

                    try:
                        entry[key] = float(v)
                    except (TypeError, ValueError):
                        continue

                if entry:
                    excel_data[tag_val] = entry

            if weight_is_tonnes and any("weight_kg" in e for e in excel_data.values()):
                self.warnings.append(
                    "Excel weight column '%s' interpreted as tonnes, converted to kg"
                    % weight_header
                )
                for entry in excel_data.values():
                    if "weight_kg" in entry:
                        entry["weight_kg"] *= 1000.0

            applied_heights = []
            for eq in equipment_list:
                tag = eq.get("tag", "")
                if tag not in excel_data:
                    continue

                ed = excel_data[tag]
                for key, val in ed.items():
                    eq[key] = round(val)
                    if key == "height_mm":
                        applied_heights.append(eq)

                eq["defaults_applied"] = [
                    k for k in eq.get("defaults_applied", [])
                    if k not in ed
                ]

                if "data_source" in eq:
                    eq["data_source"] = "svg_geometry+excel"

            if applied_heights and all(eq["height_mm"] < 50 for eq in applied_heights):
                self.warnings.append(
                    "Excel heights appear to be in metres; auto-converted to mm"
                )
                for eq in applied_heights:
                    eq["height_mm"] = int(eq["height_mm"] * 1000)

        except Exception as e:
            self.warnings.append("Excel merge failed: %s" % str(e))

    # -------------------------------------------------------------------
    # Build dimension association
    # -------------------------------------------------------------------
    def _associate_dimensions(self) -> list:
        """Associate each dimension with the features it spans."""
        result = []
        for d in self.dim_labels:
            entry = {
                "value_mm": d.value_mm,
                "direction": d.direction,
            }

            # Try to identify what features the arrow endpoints are near
            from_ref = "boundary"
            to_ref = "boundary"

            if d.direction == "horizontal":
                for s in self.shapes:
                    left = s.cx_pts - s.width_pts / 2
                    right = s.cx_pts + s.width_pts / 2
                    if s.shape_type == "ellipse":
                        left = s.cx_pts - s.rx_pts
                        right = s.cx_pts + s.rx_pts
                    if abs(left - d.end_abs) < 5:
                        to_ref = s.tag + "_left"
                    elif abs(right - d.start_abs) < 5:
                        from_ref = s.tag + "_right"
                    elif abs(s.cx_pts - d.end_abs) < 5:
                        to_ref = s.tag + "_center"
                    elif abs(s.cx_pts - d.start_abs) < 5:
                        from_ref = s.tag + "_center"

            entry["from_ref"] = from_ref
            entry["to_ref"] = to_ref

            # Real-world endpoints + perpendicular offset so the scene can
            # draw each dimension at its original location and length
            # without having to re-resolve refs.
            if self.scale:
                if d.direction == "horizontal":
                    entry["start_mm"] = round(
                        (d.start_abs - self.origin_x) * self.scale
                    )
                    entry["end_mm"] = round(
                        (d.end_abs - self.origin_x) * self.scale
                    )
                    entry["perp_mm"] = round(
                        (self.origin_y_bottom - d.perp_abs) * self.scale
                    )
                else:
                    # For vertical dims, start_abs/end_abs are SVG y values.
                    # Larger SVG y = lower on page, so y_bottom - svg_y flips.
                    s_real = (self.origin_y_bottom - d.start_abs) * self.scale
                    e_real = (self.origin_y_bottom - d.end_abs) * self.scale
                    entry["start_mm"] = round(min(s_real, e_real))
                    entry["end_mm"] = round(max(s_real, e_real))
                    entry["perp_mm"] = round(
                        (d.perp_abs - self.origin_x) * self.scale
                    )

            result.append(entry)

        return result

    # -------------------------------------------------------------------
    # Main parse method
    # -------------------------------------------------------------------
    def parse(self) -> dict:
        """Run the full parsing pipeline and return structured JSON dict."""
        # Step 1: Extract all shapes, dimensions, labels from SVG
        self._extract_shapes()

        # Step 2: Match text labels to equipment shapes
        self._match_labels_to_shapes()

        # Step 3: Pair dimension values with their arrows
        self._pair_dimensions()

        # Step 4: Calibrate scale from dimensions
        self._calibrate_scale()

        # Step 5: Detect coordinate origin
        self._detect_origin()

        # Step 6: Build equipment list in real-world coordinates
        equipment = []
        for s in self.shapes:
            eq_type = self._classify_equipment(s)
            cx_mm, cy_mm = self._to_real(s.cx_pts, s.cy_pts)

            entry = {
                "tag": s.tag,
                "svg_shape": s.shape_type,
                "equipment_type": eq_type,
                "center_x_mm": cx_mm,
                "center_y_mm": cy_mm,
                "rotation_deg": s.rotation_deg,
                "elevation_mm": DEFAULT_ELEVATION_MM,
                "data_source": "svg_geometry",
                "defaults_applied": ["height_mm", "weight_kg", "elevation_mm"],
            }

            if eq_type == "box":
                entry["width_mm"] = self._to_real_dim(s.width_pts)
                entry["depth_mm"] = self._to_real_dim(s.height_pts)
                entry["height_mm"] = DEFAULT_HEIGHT_MM
                entry["weight_kg"] = DEFAULT_WEIGHT_KG

            elif eq_type == "vertical_vessel":
                diameter = self._to_real_dim(s.rx_pts * 2)
                entry["diameter_mm"] = diameter
                entry["height_mm"] = DEFAULT_HEIGHT_MM
                entry["weight_kg"] = DEFAULT_WEIGHT_KG

            elif eq_type == "horizontal_vessel":
                # Width is the vessel length, height is the diameter
                vessel_length = self._to_real_dim(s.width_pts)
                vessel_diam = self._to_real_dim(s.height_pts)
                # Ensure length > diameter (length is the longer dimension)
                if vessel_diam > vessel_length:
                    vessel_length, vessel_diam = vessel_diam, vessel_length
                entry["length_mm"] = vessel_length
                entry["diameter_mm"] = vessel_diam
                entry["height_mm"] = DEFAULT_HEIGHT_MM
                entry["weight_kg"] = DEFAULT_WEIGHT_KG
                entry["defaults_applied"].append("saddle_height_mm")
                entry["saddle_height_mm"] = round(vessel_diam * 0.4)

            equipment.append(entry)

        # Step 7: Merge Excel data if provided
        self._merge_excel(equipment)

        # Step 8: Build grid lines
        grid = {"x_lines": [], "y_lines": []}
        for gl in self.grid_labels:
            gx, gy = self._to_real(gl.x_pts, gl.y_pts)
            if gl.label.isalpha():
                grid["y_lines"].append({"label": gl.label, "position_mm": gy})
            else:
                grid["x_lines"].append({"label": gl.label, "position_mm": gx})

        # If no grid labels found, generate from module boundary
        if not grid["x_lines"] and self.module_width_mm:
            grid["x_lines"] = [
                {"label": "1", "position_mm": 0},
                {"label": "2", "position_mm": self.module_width_mm},
            ]
            self.warnings.append(
                "No grid line labels found in SVG; generated from boundary"
            )
        if not grid["y_lines"] and self.module_length_mm:
            grid["y_lines"] = [
                {"label": "A", "position_mm": 0},
                {"label": "B", "position_mm": self.module_length_mm},
            ]

        # Step 9: Assemble result
        result = {
            "source": self.svg_path,
            "units": "mm",
            "coordinate_system": "origin bottom-left, X-right, Y-up",
            "scale_mm_per_pt": round(self.scale, 4) if self.scale else None,
            "module_boundary": {
                "width_mm": self.module_width_mm,
                "length_mm": self.module_length_mm,
                "origin_x_mm": 0,
                "origin_y_mm": 0,
            },
            "grid_lines": grid,
            "equipment": equipment,
            "dimensions": self._associate_dimensions(),
            "parser_warnings": self.warnings,
        }

        self._result = result
        return result

    def save_json(self, output_path: str):
        """Save parsed result to JSON file."""
        if not hasattr(self, "_result"):
            self.parse()
        with open(output_path, "w") as f:
            json.dump(self._result, f, indent=2)
        return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Parse Visio SVG layout to JSON")
    ap.add_argument("svg", help="Path to .svg file")
    ap.add_argument("-o", "--output", help="Output JSON path")
    ap.add_argument("--excel", help="Equipment list Excel file")
    args = ap.parse_args()

    output = args.output or args.svg.rsplit(".", 1)[0] + "_layout.json"
    parser = SvgLayoutParser(args.svg, excel_path=args.excel)
    result = parser.parse()
    parser.save_json(output)

    print("Parsed %d equipment items" % len(result["equipment"]))
    print("Scale: %.4f mm/pt" % (result["scale_mm_per_pt"] or 0))
    print("Module: %s x %s mm" % (
        result["module_boundary"]["width_mm"],
        result["module_boundary"]["length_mm"],
    ))
    if result["parser_warnings"]:
        print("\nWarnings:")
        for w in result["parser_warnings"]:
            print("  - %s" % w)
    print("\nSaved to: %s" % output)
